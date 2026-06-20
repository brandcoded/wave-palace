from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
BLUE = "0000FF"      # inputs
BLACK = "000000"     # formulas
GREEN = "008000"     # cross-sheet links
WHITE = "FFFFFF"
NAVY = "1F2A44"
ACCENT = "3B4A6B"
YELLOW = "FFF6CC"

CUR = '$#,##0;($#,##0);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'
NUM = '#,##0;(#,##0);"-"'

thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_title(cell):
    cell.font = Font(name=FONT, bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center", indent=1)

def hdr(cell):
    cell.font = Font(name=FONT, bold=True, size=10, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=ACCENT)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

def label(cell, bold=False):
    cell.font = Font(name=FONT, size=10, bold=bold, color=BLACK)
    cell.alignment = Alignment(vertical="center")

def inp(cell, fmt):
    cell.font = Font(name=FONT, size=10, color=BLUE)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    cell.border = border

def calc(cell, fmt, color=BLACK, bold=False):
    cell.font = Font(name=FONT, size=10, color=color, bold=bold)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")
    cell.border = border

# ----------------------------------------------------------------------------
# ASSUMPTIONS
# ----------------------------------------------------------------------------
a = wb.active
a.title = "Assumptions"
a.sheet_view.showGridLines = False
a.column_dimensions["A"].width = 46
a.column_dimensions["B"].width = 14
a.column_dimensions["C"].width = 52
a.merge_cells("A1:C1")
a["A1"] = "WavePalace  —  Revenue Model Assumptions"
style_title(a["A1"])
a.row_dimensions[1].height = 26
a["A2"] = "Blue / shaded cells are inputs — change these to run scenarios."
a["A2"].font = Font(name=FONT, italic=True, size=9, color="666666")

rows = [
    ("SECTION", "Audience & growth drivers", "", ""),
    ("Starting active hosts / curators (Month 1)", 50, NUM, "DJs, world owners, lounge curators with a channel"),
    ("Host base monthly growth", 0.12, PCT, "MoM growth in active hosts"),
    ("Starting monthly active listeners (Month 1)", 2000, NUM, "Web + VRChat listeners"),
    ("Listener base monthly growth", 0.15, PCT, "MoM growth in MAU"),
    ("SECTION", "Host / creator monetization", "", ""),
    ("Host Pro price ($/mo)", 12, CUR, "Branding, vanity URL, uploads, animated visuals, analytics"),
    ("Host Pro conversion (% of hosts)", 0.08, PCT, "Share of hosts on a paid Pro plan"),
    ("Featured placement price ($/mo)", 25, CUR, "Promoted slot in the directory"),
    ("Featured placement (% of hosts)", 0.05, PCT, "Share of hosts buying a featured slot"),
    ("Done-for-you channel setup ($ one-time)", 150, CUR, "Concierge: mux + host the customer's cleared media"),
    ("Done-for-you take rate (% of new hosts/mo)", 0.10, PCT, "New hosts each month buying setup"),
    ("VRChat event package ($ each)", 40, CUR, "Per-event curated set + reliable muxed link"),
    ("Event packages (% of hosts/mo)", 0.04, PCT, "Hosts running a paid event in the month"),
    ("SECTION", "Listener & passive monetization", "", ""),
    ("Tips: listeners tipping per month (% MAU)", 0.015, PCT, "Share of MAU leaving a tip"),
    ("Tips: average tip ($)", 4.00, CUR2, "Gross tip amount"),
    ("Tips: platform take rate", 0.15, PCT, "WavePalace cut of each tip"),
    ("Affiliate: listeners clicking 'Listen elsewhere' (% MAU/mo)", 0.30, PCT, "Clickthrough on existing externalLinks"),
    ("Affiliate: click-to-signup conversion", 0.04, PCT, "Share of clicks that convert"),
    ("Affiliate: commission per conversion ($)", 0.50, CUR2, "Referral payout from streaming partners"),
    ("SECTION", "B2B / white-label", "", ""),
    ("White-label client price ($/mo)", 300, CUR, "Embed player + mux pipeline for venues/communities"),
    ("White-label ramp (new clients / month)", 0.30, NUM, "Sales-led; ~1 client every ~3 months"),
]

r = 4
ASSUMP = {}  # label -> row
for lab, val, fmt, note in rows:
    if lab == "SECTION":
        a.merge_cells(f"A{r}:C{r}")
        c = a[f"A{r}"]
        c.value = val
        c.font = Font(name=FONT, bold=True, size=11, color=WHITE)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(vertical="center", indent=1)
        a.row_dimensions[r].height = 20
    else:
        label(a[f"A{r}"])
        a[f"A{r}"] = lab
        inp(a[f"B{r}"], fmt)
        a[f"B{r}"] = val
        a[f"C{r}"] = note
        a[f"C{r}"].font = Font(name=FONT, italic=True, size=9, color="888888")
        ASSUMP[lab] = r
    r += 1

# convenient absolute refs
def A(lab):
    return f"Assumptions!$B${ASSUMP[lab]}"

# ----------------------------------------------------------------------------
# PROJECTION  (12 months, months as rows)
# ----------------------------------------------------------------------------
p = wb.create_sheet("Projection")
p.sheet_view.showGridLines = False
p.merge_cells("A1:P1")
p["A1"] = "WavePalace  —  12-Month Revenue Projection"
style_title(p["A1"])
p.row_dimensions[1].height = 26

cols = [
    ("A", "Month", 9),
    ("B", "Active\nHosts", 10),
    ("C", "Active\nListeners", 11),
    ("D", "Paying\nHosts", 9),
    ("E", "Host Pro\n($/mo)", 11),
    ("F", "Featured\n($/mo)", 11),
    ("G", "White-label\n($/mo)", 11),
    ("H", "Tips net\n($/mo)", 11),
    ("I", "Affiliate\n($/mo)", 11),
    ("J", "Done-for-you\n($/mo)", 12),
    ("K", "Events\n($/mo)", 11),
    ("L", "Total Rev\n($/mo)", 12),
    ("M", "Cumulative\n($)", 13),
    ("N", "MRR\n($/mo)", 12),
    ("O", "ARPU/host\n($/mo)", 11),
    ("P", "ARPU/listener\n($/mo)", 12),
]
for col, name, width in cols:
    p.column_dimensions[col].width = width
    hdr(p[f"{col}2"])
    p[f"{col}2"] = name
p.row_dimensions[2].height = 30

first, last = 3, 14  # 12 months
for i in range(12):
    rr = first + i
    m = i + 1
    prev = rr - 1
    # month
    calc(p[f"A{rr}"], NUM)
    p[f"A{rr}"] = m
    p[f"A{rr}"].alignment = Alignment(horizontal="center")
    # hosts / listeners
    if i == 0:
        p[f"B{rr}"] = f"={A('Starting active hosts / curators (Month 1)')}"
        p[f"C{rr}"] = f"={A('Starting monthly active listeners (Month 1)')}"
    else:
        p[f"B{rr}"] = f"=B{prev}*(1+{A('Host base monthly growth')})"
        p[f"C{rr}"] = f"=C{prev}*(1+{A('Listener base monthly growth')})"
    calc(p[f"B{rr}"], NUM)
    calc(p[f"C{rr}"], NUM)
    # paying hosts
    p[f"D{rr}"] = f"=B{rr}*{A('Host Pro conversion (% of hosts)')}"
    calc(p[f"D{rr}"], NUM)
    # host pro
    p[f"E{rr}"] = f"=D{rr}*{A('Host Pro price ($/mo)')}"
    # featured
    p[f"F{rr}"] = f"=B{rr}*{A('Featured placement (% of hosts)')}*{A('Featured placement price ($/mo)')}"
    # white-label
    p[f"G{rr}"] = f"=ROUND(A{rr}*{A('White-label ramp (new clients / month)')},0)*{A('White-label client price ($/mo)')}"
    # tips
    p[f"H{rr}"] = f"=C{rr}*{A('Tips: listeners tipping per month (% MAU)')}*{A('Tips: average tip ($)')}*{A('Tips: platform take rate')}"
    # affiliate
    aff_click = A("Affiliate: listeners clicking 'Listen elsewhere' (% MAU/mo)")
    aff_conv = A("Affiliate: click-to-signup conversion")
    aff_comm = A("Affiliate: commission per conversion ($)")
    p[f"I{rr}"] = f"=C{rr}*{aff_click}*{aff_conv}*{aff_comm}"
    # done-for-you  (new hosts this month)
    if i == 0:
        newhosts = f"B{rr}"
    else:
        newhosts = f"(B{rr}-B{prev})"
    p[f"J{rr}"] = f"={newhosts}*{A('Done-for-you take rate (% of new hosts/mo)')}*{A('Done-for-you channel setup ($ one-time)')}"
    # events
    p[f"K{rr}"] = f"=B{rr}*{A('Event packages (% of hosts/mo)')}*{A('VRChat event package ($ each)')}"
    # totals
    p[f"L{rr}"] = f"=SUM(E{rr}:K{rr})"
    p[f"M{rr}"] = f"=L{rr}" if i == 0 else f"=M{prev}+L{rr}"
    p[f"N{rr}"] = f"=E{rr}+F{rr}+G{rr}+H{rr}+I{rr}"
    p[f"O{rr}"] = f"=L{rr}/B{rr}"
    p[f"P{rr}"] = f"=L{rr}/C{rr}"
    for col in "EFGHIJKLMN":
        calc(p[f"{col}{rr}"], CUR)
    calc(p[f"O{rr}"], CUR2)
    calc(p[f"P{rr}"], CUR2)
    if i % 2 == 1:
        for col in "ABCDEFGHIJKLMNOP":
            p[f"{col}{rr}"].fill = PatternFill("solid", fgColor="F4F6FA")

# totals row
tr = last + 1
p[f"A{tr}"] = "Year 1"
p[f"A{tr}"].font = Font(name=FONT, bold=True, size=10, color=WHITE)
p[f"A{tr}"].fill = PatternFill("solid", fgColor=NAVY)
p[f"A{tr}"].alignment = Alignment(horizontal="center")
for col in "EFGHIJKL":
    p[f"{col}{tr}"] = f"=SUM({col}{first}:{col}{last})"
    calc(p[f"{col}{tr}"], CUR, color=WHITE, bold=True)
    p[f"{col}{tr}"].fill = PatternFill("solid", fgColor=NAVY)
for col in "ABCDM N O P".replace(" ", ""):
    if col in ("A",):
        continue
    p[f"{col}{tr}"].fill = PatternFill("solid", fgColor=NAVY)
    p[f"{col}{tr}"].border = border
p[f"M{tr}"] = f"=M{last}"
calc(p[f"M{tr}"], CUR, color=WHITE, bold=True)
p[f"M{tr}"].fill = PatternFill("solid", fgColor=NAVY)

# ----------------------------------------------------------------------------
# SUMMARY
# ----------------------------------------------------------------------------
s = wb.create_sheet("Summary")
wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))  # move to first
s.sheet_view.showGridLines = False
for col, w in [("A", 34), ("B", 16), ("C", 14), ("D", 4), ("E", 30), ("F", 14), ("G", 12)]:
    s.column_dimensions[col].width = w
s.merge_cells("A1:G1")
s["A1"] = "WavePalace  —  Revenue Model Summary (Year 1)"
style_title(s["A1"])
s.row_dimensions[1].height = 26

# KPI block
s["A3"] = "Headline outputs"
s["A3"].font = Font(name=FONT, bold=True, size=12, color=NAVY)
kpis = [
    ("Year 1 total revenue", f"=Projection!L{tr}", CUR),
    ("Ending MRR (Month 12, recurring)", f"=Projection!N{last}", CUR),
    ("Ending total monthly revenue (Month 12)", f"=Projection!L{last}", CUR),
    ("Active hosts (Month 12)", f"=Projection!B{last}", NUM),
    ("Active listeners (Month 12)", f"=Projection!C{last}", NUM),
    ("ARPU per host / month (Month 12)", f"=Projection!O{last}", CUR2),
    ("ARPU per listener / month (Month 12)", f"=Projection!P{last}", CUR2),
]
rr = 4
for lab, f, fmt in kpis:
    label(s[f"A{rr}"], bold=True)
    s[f"A{rr}"] = lab
    s[f"B{rr}"] = f
    calc(s[f"B{rr}"], fmt, color=GREEN, bold=True)
    rr += 1

# Revenue mix table (right block)
s["E3"] = "Year 1 revenue mix by stream"
s["E3"].font = Font(name=FONT, bold=True, size=12, color=NAVY)
hdr(s["E4"]); s["E4"] = "Revenue stream"
hdr(s["F4"]); s["F4"] = "Year 1 ($)"
hdr(s["G4"]); s["G4"] = "% mix"
streams = [
    ("Host Pro subscriptions", "E"),
    ("Featured placement", "F"),
    ("White-label / B2B", "G"),
    ("Tips (net)", "H"),
    ("Affiliate", "I"),
    ("Done-for-you setup", "J"),
    ("VRChat event packages", "K"),
]
mix_first = 5
rr = mix_first
for name, col in streams:
    label(s[f"E{rr}"])
    s[f"E{rr}"] = name
    s[f"F{rr}"] = f"=Projection!{col}{tr}"
    calc(s[f"F{rr}"], CUR, color=GREEN)
    s[f"G{rr}"] = f"=F{rr}/$F${rr_total if False else mix_first+len(streams)}"
    rr += 1
mix_last = rr - 1
tot_row = rr
label(s[f"E{tot_row}"], bold=True)
s[f"E{tot_row}"] = "Total"
s[f"F{tot_row}"] = f"=SUM(F{mix_first}:F{mix_last})"
calc(s[f"F{tot_row}"], CUR, bold=True)
s[f"G{tot_row}"] = f"=SUM(G{mix_first}:G{mix_last})"
# fix % refs to total
for i in range(len(streams)):
    s[f"G{mix_first+i}"] = f"=F{mix_first+i}/$F${tot_row}"
    calc(s[f"G{mix_first+i}"], PCT)
calc(s[f"G{tot_row}"], PCT, bold=True)
for c in ("E", "F", "G"):
    s[f"{c}{tot_row}"].fill = PatternFill("solid", fgColor="F4F6FA")

# notes
nr = max(rr, 13) + 2
s[f"A{nr}"] = "Notes"
s[f"A{nr}"].font = Font(name=FONT, bold=True, size=11, color=NAVY)
notes = [
    "Drivers and pricing live on the Assumptions tab — edit the shaded cells to run scenarios.",
    "Licensing-light streams (host plans, done-for-you, white-label, affiliate, tips) are modeled first;",
    "   listener subscriptions / ad-supported radio are excluded until music licensing is secured.",
    "Done-for-you and event packages are one-time (non-recurring) and excluded from MRR.",
    "Conservative MVP-stage assumptions; not a forecast — a sizing tool for prioritization.",
]
for i, n in enumerate(notes):
    s[f"A{nr+1+i}"] = n
    s.merge_cells(f"A{nr+1+i}:G{nr+1+i}")
    s[f"A{nr+1+i}"].font = Font(name=FONT, italic=True, size=9, color="666666")

wb.calculation.fullCalcOnLoad = True
wb.save("WavePalace_Revenue_Model.xlsx")
print("saved")

# ---- independent verification of the model math (mirrors the formulas) ----
g = {k: a[f"B{ASSUMP[k]}"].value for k in ASSUMP}
hosts = g["Starting active hosts / curators (Month 1)"]
mau = g["Starting monthly active listeners (Month 1)"]
cum = 0.0
tot = {k: 0.0 for k in ["E", "F", "G", "H", "I", "J", "K", "L"]}
print(f"{'M':>2} {'Hosts':>6} {'MAU':>7} {'Pro':>6} {'Feat':>5} {'WL':>5} {'Tips':>5} {'Aff':>5} {'DFY':>5} {'Evt':>5} {'TOTAL':>7} {'Cum':>8} {'MRR':>6}")
for m in range(1, 13):
    if m > 1:
        new = hosts * g["Host base monthly growth"]
        hosts += new
        mau *= (1 + g["Listener base monthly growth"])
    else:
        new = hosts
    pro = hosts * g["Host Pro conversion (% of hosts)"] * g["Host Pro price ($/mo)"]
    feat = hosts * g["Featured placement (% of hosts)"] * g["Featured placement price ($/mo)"]
    wl = round(m * g["White-label ramp (new clients / month)"]) * g["White-label client price ($/mo)"]
    tips = mau * g["Tips: listeners tipping per month (% MAU)"] * g["Tips: average tip ($)"] * g["Tips: platform take rate"]
    aff = mau * g["Affiliate: listeners clicking 'Listen elsewhere' (% MAU/mo)"] * g["Affiliate: click-to-signup conversion"] * g["Affiliate: commission per conversion ($)"]
    dfy = new * g["Done-for-you take rate (% of new hosts/mo)"] * g["Done-for-you channel setup ($ one-time)"]
    evt = hosts * g["Event packages (% of hosts/mo)"] * g["VRChat event package ($ each)"]
    total = pro + feat + wl + tips + aff + dfy + evt
    mrr = pro + feat + wl + tips + aff
    cum += total
    for k, v in zip("EFGHIJK", [pro, feat, wl, tips, aff, dfy, evt]):
        tot[k] += v
    tot["L"] += total
    print(f"{m:>2} {hosts:>6.0f} {mau:>7.0f} {pro:>6.0f} {feat:>5.0f} {wl:>5.0f} {tips:>5.0f} {aff:>5.0f} {dfy:>5.0f} {evt:>5.0f} {total:>7.0f} {cum:>8.0f} {mrr:>6.0f}")
print("-"*90)
print(f"Year 1 total: ${tot['L']:,.0f}  | Ending MRR (M12): ${mrr:,.0f}  | ending hosts {hosts:.0f}, MAU {mau:.0f}")
print("Mix:", {k: f"{tot[k]/tot['L']:.0%}" for k in "EFGHIJK"})
